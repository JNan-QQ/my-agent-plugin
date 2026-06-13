#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
exec_json_manager.py - EXEC/BUGS JSON 管理工具

用 --type exec|bug 路由两种数据。

子命令：
    create   - 从 JSON 数据创建新文件
    read     - 读取记录（支持索引范围）
    update   - 更新指定索引的记录
    append   - 追加新记录
    to_xlsx  - 将 JSON 转换为 XLSX
"""

import argparse
import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── 样式（与 case_json_manager 一致） ──
HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT = Font(name="微软雅黑", size=10)
DATA_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# ── exec/bug schema 配置 ──
EXEC_HEADERS = ["用例编号", "执行状态", "执行时间", "执行人", "实际结果", "缺陷编号", "备注"]
EXEC_FIELDS = ["case_id", "status", "time", "executor", "actual", "bug_id", "note"]
EXEC_WIDTHS = [12, 12, 18, 12, 40, 18, 20]

BUG_HEADERS = ["缺陷编号", "缺陷标题", "严重程度", "优先级", "所属模块",
               "复现步骤", "预期结果", "实际结果", "状态", "发现时间", "关联用例"]
BUG_FIELDS = ["id", "title", "severity", "priority", "module",
              "steps", "expected", "actual", "status", "found_time", "case_id"]
BUG_WIDTHS = [18, 35, 10, 10, 15, 40, 30, 30, 10, 18, 12]


def schema(type_name):
    if type_name == "exec":
        return EXEC_HEADERS, EXEC_FIELDS, EXEC_WIDTHS
    if type_name == "bug":
        return BUG_HEADERS, BUG_FIELDS, BUG_WIDTHS
    raise SystemExit(f"Unknown --type: {type_name}")


def load_json(path):
    return json.loads(Path(path).read_text(encoding="utf-8"))


def save_json(path, data):
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    Path(path).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def cmd_create(args):
    data = load_json(args.data)
    if not isinstance(data, list):
        data = [data]
    save_json(args.output, data)
    print(f"Created: {args.output} ({len(data)} records, type={args.type})", file=sys.stderr)


def cmd_read(args):
    records = load_json(args.file)
    start = args.start if args.start is not None else 0
    end = args.end if args.end is not None else len(records)
    print(json.dumps(records[start:end], ensure_ascii=False, indent=2))


def cmd_append(args):
    records = load_json(args.file)
    new_records = load_json(args.data)
    if not isinstance(new_records, list):
        new_records = [new_records]
    records.extend(new_records)
    save_json(args.file, records)
    print(f"Appended {len(new_records)} records (type={args.type})", file=sys.stderr)


def cmd_update(args):
    records = load_json(args.file)
    update_data = load_json(args.data)
    if args.index >= len(records):
        print(f"Error: index {args.index} out of range", file=sys.stderr)
        sys.exit(1)
    records[args.index].update(update_data)
    save_json(args.file, records)
    print(f"Updated record at index {args.index} (type={args.type})", file=sys.stderr)


def cmd_to_xlsx(args):
    records = load_json(args.file)
    headers, fields, widths = schema(args.type)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "执行记录" if args.type == "exec" else "缺陷记录"

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    for row_idx, rec in enumerate(records, 2):
        for col_idx, key in enumerate(fields, 1):
            ws.cell(row=row_idx, column=col_idx, value=rec.get(key, ""))
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT
            cell.border = THIN_BORDER

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{len(records) + 1}"

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    wb.close()
    print(f"Converted to XLSX: {output_path} (type={args.type})", file=sys.stderr)


def main():
    if sys.platform == "win32":
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")

    parser = argparse.ArgumentParser(description="EXEC/BUGS JSON 管理工具")
    sub = parser.add_subparsers(dest="command", required=True)

    def add_type(p):
        p.add_argument("--type", required=True, choices=["exec", "bug"])

    p_create = sub.add_parser("create")
    add_type(p_create)
    p_create.add_argument("--output", required=True)
    p_create.add_argument("--data", required=True)

    p_read = sub.add_parser("read")
    add_type(p_read)
    p_read.add_argument("--file", required=True)
    p_read.add_argument("--start", type=int, default=None)
    p_read.add_argument("--end", type=int, default=None)

    p_append = sub.add_parser("append")
    add_type(p_append)
    p_append.add_argument("--file", required=True)
    p_append.add_argument("--data", required=True)

    p_update = sub.add_parser("update")
    add_type(p_update)
    p_update.add_argument("--file", required=True)
    p_update.add_argument("--index", type=int, required=True)
    p_update.add_argument("--data", required=True)

    p_to_xlsx = sub.add_parser("to_xlsx")
    add_type(p_to_xlsx)
    p_to_xlsx.add_argument("--file", required=True)
    p_to_xlsx.add_argument("--output", required=True)

    args = parser.parse_args()
    {
        "create": cmd_create,
        "read": cmd_read,
        "append": cmd_append,
        "update": cmd_update,
        "to_xlsx": cmd_to_xlsx,
    }[args.command](args)


if __name__ == "__main__":
    main()
