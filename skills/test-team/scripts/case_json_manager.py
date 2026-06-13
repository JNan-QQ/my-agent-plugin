#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
case_json_manager.py - JSON 格式测试用例管理工具

子命令：
    create   - 从 JSON 数据创建新文件
    read     - 读取用例（支持索引范围）
    update   - 更新指定索引的用例
    append   - 追加新用例
    delete   - 删除指定索引的用例
    to_xlsx  - 将 JSON 转换为 XLSX

用法：
    python case_json_manager.py create --output cases.json --data input.json
    python case_json_manager.py read --file cases.json --start 0 --end 10
    python case_json_manager.py update --file cases.json --index 0 --data update.json
    python case_json_manager.py append --file cases.json --data new_cases.json
    python case_json_manager.py delete --file cases.json --indices 0,1,2
    python case_json_manager.py to_xlsx --file cases.json --output output.xlsx
"""

import argparse
import json
import sys
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── 样式定义 ──────────────────────────────────────────────

HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT = Font(name="微软雅黑", size=10)
DATA_ALIGNMENT = Alignment(vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

PRIORITY_FILLS = {
    "高": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
    "中": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
    "低": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
}

HEADERS = ["用例编号", "用例标题", "前置条件", "测试步骤", "预期结果",
           "优先级", "用例类型", "关联需求", "备注"]
COL_WIDTHS = [12, 35, 30, 50, 45, 10, 12, 12, 20]
FIELD_KEYS = ["id", "title", "precondition", "steps", "expected",
              "priority", "type", "req", "note"]


def load_json(path):
    """加载 JSON 文件。"""
    return json.loads(Path(path).read_text(encoding="utf-8"))


def save_json(path, data):
    """保存 JSON 文件。"""
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    Path(path).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def cmd_create(args):
    """创建新 JSON 文件。"""
    data = load_json(args.data)
    if not isinstance(data, list):
        data = [data]

    # 添加索引
    for i, case in enumerate(data):
        case["index"] = i

    save_json(args.output, data)
    print(f"Created: {args.output} ({len(data)} cases)", file=sys.stderr)


def cmd_read(args):
    """读取用例（支持索引范围）。"""
    cases = load_json(args.file)

    start = args.start if args.start is not None else 0
    end = args.end if args.end is not None else len(cases)

    result = cases[start:end]
    print(json.dumps(result, ensure_ascii=False, indent=2))


def cmd_update(args):
    """更新指定索引的用例。"""
    cases = load_json(args.file)
    update_data = load_json(args.data)

    if args.index >= len(cases):
        print(f"Error: index {args.index} out of range", file=sys.stderr)
        sys.exit(1)

    cases[args.index].update(update_data)
    save_json(args.file, cases)
    print(f"Updated case at index {args.index}", file=sys.stderr)


def cmd_append(args):
    """追加新用例。"""
    cases = load_json(args.file)
    new_cases = load_json(args.data)

    if not isinstance(new_cases, list):
        new_cases = [new_cases]

    start_idx = len(cases)
    for i, case in enumerate(new_cases):
        case["index"] = start_idx + i

    cases.extend(new_cases)
    save_json(args.file, cases)
    print(f"Appended {len(new_cases)} cases", file=sys.stderr)


def cmd_delete(args):
    """删除指定索引的用例。"""
    cases = load_json(args.file)
    indices = sorted([int(x) for x in args.indices.split(",")], reverse=True)

    for idx in indices:
        if 0 <= idx < len(cases):
            del cases[idx]

    # 重新编号索引
    for i, case in enumerate(cases):
        case["index"] = i

    save_json(args.file, cases)
    print(f"Deleted {len(indices)} cases", file=sys.stderr)


def cmd_to_xlsx(args):
    """将 JSON 转换为 XLSX。"""
    cases = load_json(args.file)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # 写入表头
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # 设置列宽
    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 写入数据
    for row_idx, case in enumerate(cases, 2):
        for col_idx, key in enumerate(FIELD_KEYS, 1):
            value = case.get(key, "")
            ws.cell(row=row_idx, column=col_idx, value=value)

        # 应用样式
        priority = case.get("priority", "").strip()
        fill = PRIORITY_FILLS.get(priority)
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill

    # 冻结���行、设置筛选
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{len(cases) + 1}"

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    wb.close()
    print(f"Converted to XLSX: {output_path}", file=sys.stderr)


# ── analyze 实现 ─────────────────────────────────────

ABNORMAL_KEYWORDS = [
    "异常", "失败", "不可用", "无效", "错误", "缺失", "冲突", "中断",
    "超时", "损坏", "不足", "不收敛", "低置信", "退出", "降级",
    "不存在", "为空", "不兼容", "不匹配", "拒绝", "过期", "重复",
    "丢失", "溢出", "不完整", "阻塞", "崩溃", "断开", "拦截"
]

BOUNDARY_KEYWORDS = ["边界", "阈值", "恰好", "刚好", "最小", "最大", "临界", "上限", "下限"]

# 质量与封驳阈值（与原 xlsx_analyzer 一致）
MAX_REASONABLE_STEPS = 10          # 用例步骤数上限
MIN_EXPECTED_LENGTH = 10           # 预期结果最短字符数
QUALITY_ISSUE_THRESHOLD = 5        # 质量问题数封驳阈值
MAX_MISSING_ABNORMAL_RATIO = 0.5   # 缺失异常用例的需求占比上限
REJECT_REQ_COVERAGE = 85           # 需求覆盖率封驳线 (%)
PASS_REQ_COVERAGE = 95             # 需求覆盖率通过线 (%)
PASS_SCENE_COVERAGE = 70           # 场景覆盖率通过线 (%)
NUM_SCENE_TYPES = 3                # 正常 / 异常 / 边界


def parse_req_range(req_str):
    """F001-F054 → ['F001', ..., 'F054']；逗号分隔也支持。"""
    match = re.match(r"([A-Z]+)(\d+)-\1(\d+)", req_str)
    if match:
        prefix = match.group(1)
        start = int(match.group(2))
        end = int(match.group(3))
        return [f"{prefix}{i:03d}" for i in range(start, end + 1)]
    return [r.strip() for r in req_str.split(",") if r.strip()]


def classify_scene(case):
    title = case.get("title", "")
    case_type = case.get("type", "")
    note = case.get("note", "")
    if any(kw in title for kw in BOUNDARY_KEYWORDS) or "边界值" in note:
        return "boundary"
    if case_type == "异常测试" or any(kw in title for kw in ABNORMAL_KEYWORDS):
        return "abnormal"
    return "normal"


def analyze_cases(cases, req_list, p0_list):
    result = {}

    # 1. 需求覆盖
    req_coverage = {r: {"cases": [], "count": 0} for r in req_list}
    for case in cases:
        req = case.get("req", "").strip()
        if req in req_coverage:
            req_coverage[req]["cases"].append(case.get("id", ""))
            req_coverage[req]["count"] += 1
    uncovered = [r for r in req_list if req_coverage[r]["count"] == 0]
    covered_count = sum(1 for r in req_list if req_coverage[r]["count"] > 0)
    req_rate = round((covered_count / len(req_list) * 100), 1) if req_list else 0.0
    result["requirement_coverage"] = req_coverage
    result["requirement_coverage_rate"] = req_rate
    result["uncovered_requirements"] = uncovered

    # 2. 场景覆盖
    scene_cov = {}
    for r in req_list:
        scene_cov[r] = {"normal": False, "abnormal": False, "boundary": False,
                        "normal_cases": [], "abnormal_cases": [], "boundary_cases": []}
    for case in cases:
        req = case.get("req", "").strip()
        if req not in scene_cov:
            continue
        scene = classify_scene(case)
        scene_cov[req][scene] = True
        scene_cov[req][f"{scene}_cases"].append(case.get("id", ""))

    total_scenes = len(req_list) * NUM_SCENE_TYPES
    covered_scenes = 0
    normal_covered = abnormal_covered = boundary_covered = 0
    missing_abnormal = []
    for r in req_list:
        if scene_cov[r]["normal"]:
            covered_scenes += 1
            normal_covered += 1
        if scene_cov[r]["abnormal"]:
            covered_scenes += 1
            abnormal_covered += 1
        else:
            missing_abnormal.append(r)
        if scene_cov[r]["boundary"]:
            covered_scenes += 1
            boundary_covered += 1
    scene_rate = round((covered_scenes / total_scenes * 100), 1) if total_scenes else 0

    result["scene_coverage"] = scene_cov
    result["scene_coverage_rate"] = scene_rate
    result["normal_coverage"] = f"{normal_covered}/{len(req_list)}"
    result["abnormal_coverage"] = f"{abnormal_covered}/{len(req_list)}"
    result["boundary_coverage"] = f"{boundary_covered}/{len(req_list)}"
    result["missing_abnormal_requirements"] = missing_abnormal

    # 3. P0 深度
    p0_single = []
    for r in p0_list:
        if r in req_coverage and req_coverage[r]["count"] <= 1:
            p0_single.append(r)
    result["p0_single_case_requirements"] = p0_single

    # 4. 质量检查
    quality = {"total_cases": len(cases), "long_steps": [],
               "short_expected": [], "duplicate_titles": [], "numbering_issues": []}
    for case in cases:
        steps = case.get("steps", "")
        step_count = len(re.findall(r"^\d+\.", steps, re.MULTILINE))
        if not step_count:
            step_count = len(steps.split("\n"))
        if step_count > MAX_REASONABLE_STEPS:
            quality["long_steps"].append(case.get("id", ""))
    for case in cases:
        if len(case.get("expected", "")) < MIN_EXPECTED_LENGTH:
            quality["short_expected"].append(case.get("id", ""))
    title_map = {}
    for case in cases:
        t = case.get("title", "").strip()
        if t:
            title_map.setdefault(t, []).append(case.get("id", ""))
    for t, ids in title_map.items():
        if len(ids) > 1:
            quality["duplicate_titles"].append({"title": t, "cases": ids})
    nums = []
    for case in cases:
        m = re.match(r"TC(\d+)", case.get("id", ""))
        if m:
            nums.append(int(m.group(1)))
    nums.sort()
    for i in range(1, len(nums)):
        if nums[i] != nums[i-1] + 1:
            quality["numbering_issues"].append(f"Gap between TC{nums[i-1]:03d} and TC{nums[i]:03d}")
    result["quality_checks"] = quality

    # 5. 分布
    pri_dist, type_dist = {}, {}
    for case in cases:
        p = case.get("priority", "").strip()
        t = case.get("type", "").strip()
        if p:
            pri_dist[p] = pri_dist.get(p, 0) + 1
        if t:
            type_dist[t] = type_dist.get(t, 0) + 1
    result["priority_distribution"] = pri_dist
    result["type_distribution"] = type_dist

    # 6. 封驳判定
    reject = []
    cond = []
    if req_rate < REJECT_REQ_COVERAGE:
        reject.append(f"Requirement coverage {req_rate}% < {REJECT_REQ_COVERAGE}%")
    if uncovered:
        reject.append(f"Uncovered requirements: {uncovered}")
    if len(missing_abnormal) > len(req_list) * MAX_MISSING_ABNORMAL_RATIO:
        reject.append(f"{len(missing_abnormal)}/{len(req_list)} requirements missing abnormal cases")
    if p0_single:
        reject.append(f"P0 requirements with single case: {p0_single}")
    if len(quality["long_steps"]) > QUALITY_ISSUE_THRESHOLD:
        reject.append(f"{len(quality['long_steps'])} cases with >10 steps")
    if len(quality["short_expected"]) > QUALITY_ISSUE_THRESHOLD:
        reject.append(f"{len(quality['short_expected'])} cases with vague expected results")

    if reject:
        verdict = "REJECT"
    elif req_rate >= PASS_REQ_COVERAGE and not p0_single:
        if scene_rate >= PASS_SCENE_COVERAGE:
            verdict = "PASS"
        else:
            verdict = "CONDITIONAL"
            cond.append(f"Scene coverage {scene_rate}% (consider improving)")
    else:
        verdict = "CONDITIONAL"
        if req_rate < PASS_REQ_COVERAGE:
            cond.append(f"Requirement coverage {req_rate}% < {PASS_REQ_COVERAGE}%")

    result["verdict"] = {"result": verdict, "reject_reasons": reject, "conditional_notes": cond}
    return result


def cmd_analyze(args):
    cases = load_json(args.file)
    req_list = parse_req_range(args.reqs)
    p0_list = [r.strip() for r in args.p0.split(",") if r.strip()] if args.p0 else []
    result = analyze_cases(cases, req_list, p0_list)
    output = json.dumps(result, ensure_ascii=False, indent=2)
    if args.output:
        Path(args.output).write_text(output, encoding="utf-8")
        print(f"Analysis saved to: {args.output}", file=sys.stderr)
    else:
        print(output)


def main():
    if sys.platform == "win32":
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")

    parser = argparse.ArgumentParser(description="JSON 格式测试用例管理工具")
    sub = parser.add_subparsers(dest="command", required=True)

    # create
    p_create = sub.add_parser("create", help="创建新 JSON 文件")
    p_create.add_argument("--output", required=True, help="输出文件路径")
    p_create.add_argument("--data", required=True, help="输入 JSON 数据文件")

    # read
    p_read = sub.add_parser("read", help="读取用例")
    p_read.add_argument("--file", required=True, help="JSON 文件路径")
    p_read.add_argument("--start", type=int, default=None, help="起始索引")
    p_read.add_argument("--end", type=int, default=None, help="结束索引")

    # update
    p_update = sub.add_parser("update", help="更新指定索引的用例")
    p_update.add_argument("--file", required=True, help="JSON 文件路径")
    p_update.add_argument("--index", type=int, required=True, help="用例索引")
    p_update.add_argument("--data", required=True, help="更新数据 JSON 文件")

    # append
    p_append = sub.add_parser("append", help="追加新用例")
    p_append.add_argument("--file", required=True, help="JSON 文件路径")
    p_append.add_argument("--data", required=True, help="新用例 JSON 文件")

    # delete
    p_delete = sub.add_parser("delete", help="删除指定索引的用例")
    p_delete.add_argument("--file", required=True, help="JSON 文件路径")
    p_delete.add_argument("--indices", required=True, help="索引列表（逗号分隔）")

    # to_xlsx
    p_to_xlsx = sub.add_parser("to_xlsx", help="转换为 XLSX")
    p_to_xlsx.add_argument("--file", required=True, help="JSON 文件路径")
    p_to_xlsx.add_argument("--output", required=True, help="输出 XLSX 文件路径")

    # analyze
    p_analyze = sub.add_parser("analyze", help="覆盖度+质量分析")
    p_analyze.add_argument("--file", required=True, help="JSON 用例文件")
    p_analyze.add_argument("--reqs", required=True, help="需求范围（如 F001-F054）")
    p_analyze.add_argument("--p0", default="", help="P0 需求列表（逗号分隔）")
    p_analyze.add_argument("--output", default=None, help="输出 JSON 路径（默认 stdout）")

    args = parser.parse_args()

    if args.command == "create":
        cmd_create(args)
    elif args.command == "read":
        cmd_read(args)
    elif args.command == "update":
        cmd_update(args)
    elif args.command == "append":
        cmd_append(args)
    elif args.command == "delete":
        cmd_delete(args)
    elif args.command == "to_xlsx":
        cmd_to_xlsx(args)
    elif args.command == "analyze":
        cmd_analyze(args)


if __name__ == "__main__":
    main()
